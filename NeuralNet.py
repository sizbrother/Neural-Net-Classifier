"""
Filename: NeuralNet
Author: Zaman Khan
Data Last Modified: 12/08/2022

------------------------------------------------------------------------------------------------------

Description: Implementation of neural networks with zero, one, or more hidden layers to perform
             binary classifaction, multiclass classification, and even non-classifaction problems.


    1) NeuralNetwork - A Neural Network Class
       
        a. get_outputs - Retrieves the outputs of the neural network

        b. present_info - Used for testing, presents information about the object

        c. forward_propagate - Forward propagates based on inputs. Makes predictions and 
                               updates activation values

        d. back_propagate - Back propagates. Updates error values and weights

        e. stop_learning - Determines if the learning is complete

        f. back_propagation_learning - Trains the network using back-propagation learning algorithm

    2) repetitive_experiment - Performs repetitive experiments to train models and saves the results 
                               to excel

    3) k_fold_cross_validation - Performs k-fold cross validation for assessing the accuracy
                                 of a neural network model

    4) binary_classification_mean_squared_error_accuracy - Computes the accuracy of a network on given pairs as:
                                                           accuracy = 1 - mean squared error

    5) multi_non_classification_mean_squared_error_accuracy - Computes the accuracy of a network on given pairs as:
                                                        accuracy = 1 - mean squared error

    6) read_data - Reads datafile using given delimiter

    7) convert_data_to_pairs - Turns a data list of lists into a list of (attribute, target) pairs
    
    Uncomment corresponding blocks of code in main function to work with different problems. To work with
    non-classification problems or multi-class classification problems, change the code in NeuralNetwork.stop_learning(),
    repetitive_experiment(), and k_fold_cross_validation() as well.
------------------------------------------------------------------------------------------------------


Usage: python3 project3.py DATASET.csv
"""

import csv, sys, random, math

#CITE: https://stackoverflow.com/questions/40385689/add-a-new-sheet-to-a-existing-workbook-in-python
#HELP: learned to work with excel sheets with openpyxl
from openpyxl import load_workbook

def read_data(filename, delimiter=",", has_header=True):
    """Reads datafile using given delimiter. Returns a header and a list of
    the rows of data."""
    data = []
    header = []
    with open(filename) as f:
        reader = csv.reader(f, delimiter=delimiter)
        if has_header:
            header = next(reader, None)
        for line in reader:
            example = [float(x) for x in line]
            data.append(example)

        return header, data

def convert_data_to_pairs(data, header):
    """Turns a data list of lists into a list of (attribute, target) pairs."""
    pairs = []
    for example in data:
        x = []
        y = []
        for i, element in enumerate(example):
            if header[i].startswith("target"):
                y.append(element)
            else:
                x.append(element)
        pair = (x, y)
        pairs.append(pair)
    return pairs

def dot_product(v1, v2):
    """Computes the dot product of v1 and v2"""
    sum = 0
    for i in range(len(v1)):
        sum += v1[i] * v2[i]
    return sum

def logistic(x):
    """Logistic / sigmoid function"""
    try:
        denom = (1 + math.e ** -x)
    except OverflowError:
        return 0.0
    return 1.0 / denom

def binary_classification_mean_squared_error_accuracy(nn, pairs):
    """Computes the accuracy of a network on given pairs as:
          accuracy = 1 - mean squared error
    Assumes nn has a forward_propagate method that takes an input vector x as
    its parameter, and that it has a get_outputs method that returns the
    output vector y. Since this is for binary classification, y should
    only contain a single value for use here (but could include the whole
    output vector for multiclass problems that can't have their accuracy checked
    by this function).
    
    Note: this will not work for non-classification problems like the 3-bit
    incrementer. You will need to write a different accuracy function for 
    multi-class problems."""

    error = 0
    total = len(pairs)

    for (x, y) in pairs:
        nn.forward_propagate(x)
        outputs = nn.get_outputs()
        error += (outputs[0] - y[0]) ** 2
        
        ## If you uncomment the following line, you will see the correct output
        ## and predicted output for each input.
        #print("x =", x, ", y =", y, ", outputs =", outputs)

    return 1 - (error / total)


def multi_non_classification_mean_squared_error_accuracy(nn, pairs):
    """Computes the accuracy of a network on given pairs as:
          accuracy = 1 - mean squared error, where error from one of the inputs counts as one error"""
    error = 0
    total = len(pairs)

    for (x, y) in pairs:
        nn.forward_propagate(x)
        outputs = nn.get_outputs()
        
        current_error = 0
        output_num = len(outputs)
        for index in range(output_num):
            current_error += outputs[index] - y[index]
        error += current_error ** 2
        
        ## If you uncomment the following line, you will see the correct output
        ## and predicted output for each input.
        print("x =", x, ", y =", y, ", outputs =", outputs)
    return 1 - (error / total / output_num)

################################################################################
### Neural Network code goes here

class NeuralNetwork:
    def __init__(self, layer_info):
        """Constructor for NeuralNetwork class

        Args:
            layer_info (int list): A list telling how many layers there are and how many nodes are in each layer.
                                   For instance, [6, 3, 5, 2] represents input layer with 6 nodes, two hidden layers
                                   with 3 and 5 nodes respectively, and output with 2 nodes.
        """
            
        #Initialize node index as 1
        node_index = 0
        #Keeps track of which nodes are in each layer
        self._layers = []
        
        #Initialize _layers to contain which nodes are in each layer
        for layer in layer_info:
            current_layer = []
            for _ in range(layer):
                node_index += 1
                current_layer.append(node_index)
            self._layers.append(current_layer)
        
        #Store the total number of nodes in the network
        self._node_num = node_index
        #_activations[i] is a_i; a_0 = 1
        self._activations = [1]
        for _ in range(node_index): self._activations.append(None)      
        #Weighted sum of inputs; _sum_inputs[i] is in_i;
        self._sum_inputs = [None for _ in range(node_index + 1)]        
        #_node_errors[i] is delta[i]
        self._node_errors = [0 for _ in range(node_index + 1)]
        #A list of dictionaries; _weights[i][j] is W_(i,j)
        self._weights = [{} for _ in range(node_index + 1)]
            
        #Keeps track of the number of epochs the learning process has been through
        self._epoch_num = 0
        
    
    def get_outputs(self):
        """Gets the outputs of the neural network

           Returns: (float list) activation values in the output layer
        """
        outputs = []
        for output_index in self._layers[-1]:
            outputs.append(self._activations[output_index])
        return outputs
    
    def present_info(self):
        """Presents all information of the object. Did not implement __str__ or __repr__ in order to
           better fit different data types. Mainly used for testing purposes"""
        print("The nodes in the Neural Networks are:", self._layers)
        print("The activation values are:", self._activations)
        print("The weighted sum of inputs are:", self._sum_inputs)
        print("The errors of nodes are:", self._node_errors)
        print("The weights in the network are:", self._weights)
    
    def forward_propagate(self, current_input):
        """Forward propagates based on inputs; makes predictions and updates activation values
           Args:
               current_input (float list): A list of all inputs;  
        """            
        #Activation values in the input layer are just the input values
        for input_node in self._layers[0]:
            self._activations[input_node] = current_input[input_node]
        #Loops through every layer that is not the input layer
        for layer_index in range(1, len(self._layers)):
            current_layer = self._layers[layer_index]
            #Makes predictins for each node, and updates activation values
            for j in current_layer:
                prediction = 0
                for i in range(len(self._weights)):
                    if j in self._weights[i]:
                        prediction += self._weights[i][j] * self._activations[i]
                self._sum_inputs[j] = prediction
                self._activations[j] = logistic(prediction)
    
    def back_propagate(self, training_example):
        """Back propagates; updates error values and weights
           Args:
               training_example (float list tuple): A single training example; ([inputs], [outputs])
        """
        #Gets the current errors of the output layer
        for output_index in range(len(self._layers[-1])):
            a_j = self._activations[self._layers[-1][output_index]]
            self._node_errors[self._layers[-1][output_index]] = a_j * (1 - a_j) * (training_example[1][output_index] - a_j)
        
        #Back-propagates in layers that are neither the output layer nor the input layer
        for layer_index in range(len(self._layers) - 2, 0, -1):
            current_layer = self._layers[layer_index]
            for i in current_layer:
                #Calculates the sum of all weights multiplies errors of respective nodes
                sum_weight_error = 0
                for j in self._weights[i]:
                    sum_weight_error += self._weights[i][j] * self._node_errors[j]
                a_i = self._activations[i]
                #Updates errors of each node
                self._node_errors[i] = a_i * (1 - a_i) * sum_weight_error
        
        #Change alpha value here
        alpha = 1000 / (1000 + self._epoch_num)
        # alpha = 1
        #Updates weights
        for i in range(len(self._weights)):
            for j in self._weights[i]:
                self._weights[i][j] += alpha * self._activations[i] * self._node_errors[j]
                
    def stop_learning(self, training_data):
        """Tells if the learning has been completed and should be stopped.
           Args:
               training_data (float list tuple list): All training examples
           Returns:
               True if learning has been completed; false otherwise"""
        #Stop learning after a certain number of epochs
        return self._epoch_num >= 300
        
        #Stop learning after the network has reached a certain accuracy
        # return binary_classification_mean_squared_error_accuracy(self, training_data) > 0.95
    
        #Stop learning after the network has attained a certain accuracy,
        #or has reached a certain number of epochs
        return binary_classification_mean_squared_error_accuracy(self, training_data) > 0.99 or self._epoch_num >= 1000
    
        #For non-classification problems or multi-class problems
        # return multi_non_classification_mean_squared_error_accuracy(self, training_data) > 0.99 or self._epoch_num >= 1000

    def back_propagation_learning(self, training_data):
        """Trains the network using back-propagation learning algorithm.
           
           Args:
               training_data (float list tuple list): All training examples
        """
        self._epoch_num = 0
        #Initialize weights that have dummy variable
        for node_index in range(self._node_num):
            self._weights[0][node_index + 1] = random.uniform(-1, 1)
            
        #Initialize weights irrelevant to dummy variables
        layer_1 = None
        for layer_2 in self._layers:
            if(layer_1):
                for node_1 in layer_1:
                    for node_2 in layer_2:
                        self._weights[node_1][node_2] = random.uniform(-1, 1)
            layer_1 = layer_2
            
        #Start learning
        while not self.stop_learning(training_data):
            for example in training_data:
                self.forward_propagate(example[0])
                self.back_propagate(example)
            self._epoch_num += 1
                
        
def repetitive_experiment(problem_label, layer_info, training_data, experiment_num):
    """Do repetitive experiments training models and save the results (epochs stopped and accuracy) to excel
       Args:
           problem_label (string): the name of the problem to experiment on.
           layer_info (int list): A list telling how many layers there are and how many nodes are in each layer;
                                  used for constructing and training neural network
           training_data (float list tuple list): All training examples
           experiment_num (int): the number of models trained and experimented on.
    """
    #CITE: https://openpyxl.readthedocs.io/en/stable/performance.html
    #HELP: Helped me learn to use openpyxl
    record = load_workbook("results.xlsx")
    results = record.create_sheet(problem_label)
    
    for i in range(experiment_num):
        #Creates and trains the network
        nn = NeuralNetwork(layer_info)
        nn.back_propagation_learning(training_data)
        
        #Test the ending accuracy and add it to an excel sheet
        accuracy = binary_classification_mean_squared_error_accuracy(nn, training_data)
        #Change to the following line to work with non_classification problems or multi-class problems
        # accuracy = multi_non_classification_mean_squared_error_accuracy(nn, training_data)
        print("accuracy =", accuracy)
        epoch_num = nn._epoch_num
        print("epoch num=", epoch_num)
        
        results.append([layer_info.__str__(), accuracy, epoch_num])
    
    #Save the results to the excel file
    record.save("results.xlsx")

def k_fold_cross_validation(data, neural_network, k = 5):
    """ Performs k-fold cross validation for assessing the accuracy
        of a neural network model.
        Args:
            data = A list containing the training data 
            neural_network = neural network model
            k = number of folds to be performed; default value set to 5
    """

    total_accuracy = 0

    # Randomly reorganizes the content of our data
    random.shuffle(data)

    # Repeat the process k-times
    for i in range(k):

        # Each fold will have this much data
        fold_number = int(len(data) / k)

        # Prepare the training and test data corresponding to the given k value
        training_data = data[:fold_number*i] + data[(fold_number*i)+fold_number:]
        test_data = data[fold_number*i:(fold_number*i)+fold_number]

        # Train and test the model
        neural_network.back_propagation_learning(training_data)
        accuracy = binary_classification_mean_squared_error_accuracy(neural_network, test_data)
        #Change to the following line to work with non_classification problems or multi-class problems
        # accuracy = multi_non_classification_mean_squared_error_accuracy(neural_network, test_data)
        total_accuracy += accuracy

        # Uncomment the following line to display the accuracy of each k-fold
        print("Accuracy =", accuracy)

    # Summaraizes the accuracy of the model by averaging the results
    print("The Average Accuracy = " + str(total_accuracy / k))


def main():
    print(sys.argv[0])
    header, data = read_data(sys.argv[1], ",")

    pairs = convert_data_to_pairs(data, header)

    #Binary classification
    #nn = NeuralNetwork([2, 1])
    # Note: add 1.0 to the front of each x vector to account for the dummy input
    #training = [([1.0] + x, y) for (x, y) in pairs]
    #training = [(x, [1 if y[0] - 1 == specific_class else 0 for specific_class in range(class_count)]) for (x, y) in training]

    #k_fold_cross_validation(training, [13, 20, 3])

    #Multiclass wine problem
    class_count = 3
    nn = NeuralNetwork([13, 5, class_count])
    training = [([1.0] + x, y) for (x, y) in pairs]
    training = [(x, [1 if y[0] - 1 == specific_class else 0 for specific_class in range(class_count)]) for (x, y) in training]
    k_fold_cross_validation(training, nn)


    #None-classification 3-bit increment problem
#     nn = NeuralNetwork([3, 10, 3])
#     training = [([1.0] + x, y) for (x, y) in pairs]
    
    # Check out the data:  
#     for example in training:
#         print(example)

    #k_fold_cross_validation(training, nn)
      
    nn.back_propagation_learning(training)
    print("Accuracy =", binary_classification_mean_squared_error_accuracy(nn, training))
    #For non-classification problems and multi-class problems
    # print("Accuracy =", multi_non_classification_mean_squared_error_accuracy(nn, training))
    
    print("Epoch Count = ", nn._epoch_num)
    
    #repetitive_experiment("Multi-Classification", [13, 10, 3], training, 20)
    

if __name__ == "__main__":
    main()
